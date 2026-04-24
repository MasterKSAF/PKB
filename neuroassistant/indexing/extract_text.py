"""Text extraction for indexing (PDF/XLSX)."""

from __future__ import annotations

from pathlib import Path

import fitz  # pymupdf
from openpyxl import load_workbook


def extract_text(path: Path, max_chars: int = 200_000) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf(path, max_chars=max_chars)
    if suffix == ".xlsx":
        return _extract_xlsx(path, max_chars=max_chars)
    # Fallback: not supported in MVP
    return ""


def _extract_pdf(path: Path, max_chars: int) -> str:
    doc = fitz.open(str(path))
    parts: list[str] = []
    try:
        for i in range(doc.page_count):
            txt = doc.load_page(i).get_text("text") or ""
            if txt:
                parts.append(txt)
            if sum(len(p) for p in parts) >= max_chars:
                break
        return "\n".join(parts)[:max_chars]
    finally:
        doc.close()


def _extract_xlsx(path: Path, max_chars: int) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"[SHEET] {sheet}")
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    parts.append(" | ".join(vals))
                if sum(len(p) for p in parts) >= max_chars:
                    break
        return "\n".join(parts)[:max_chars]
    finally:
        wb.close()

