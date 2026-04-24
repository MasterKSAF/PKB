"""Endpoints helpers for browsing locally downloaded datasets."""

from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class DatasetConfig:
    raw_root: Path

    @classmethod
    def from_env(cls) -> "DatasetConfig":
        root = os.environ.get("RAW_ROOT", "./data_raw")
        return cls(raw_root=Path(root))


def list_drive_unzipped_files(cfg: DatasetConfig) -> list[Path]:
    base = (cfg.raw_root / "drive_dataset_unzipped").resolve()
    if not base.exists():
        return []
    return sorted([p for p in base.rglob("*") if p.is_file()])


def read_xlsx_preview(path: Path, max_rows: int = 200) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out: dict = {"filename": path.name, "sheets": {}}
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[list[str | int | float | None]] = []
            read_rows = 0
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                rows.append(list(row))
                read_rows = i
                if i >= max_rows:
                    break
            max_row = ws.max_row or 0
            max_column = ws.max_column or 0
            out["sheets"][name] = {
                "rows": rows,
                "truncated": max_row > read_rows,
                "max_row": max_row,
                "max_column": max_column,
            }
        return out
    finally:
        wb.close()

